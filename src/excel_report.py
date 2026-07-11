"""
Excel Report Generator for Credit Risk Scoring System
Generates comprehensive Excel reports with pivot tables and charts
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

def load_data(cleaned_data_path, risk_scores_path):
    """Load cleaned data and risk scores"""
    print("Loading data...")
    
    cleaned_df = pd.read_csv(cleaned_data_path)
    risk_scores_df = pd.read_csv(risk_scores_path)
    
    print("Data loaded successfully")
    return cleaned_df, risk_scores_df

def create_kpi_sheet(wb, cleaned_df, risk_scores_df):
    """Create KPI summary sheet"""
    print("Creating KPI sheet...")
    
    ws = wb.create_sheet("KPI Summary", 0)
    
    # Title
    ws['A1'] = "Credit Risk Scoring System - KPI Summary"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:B1')
    
    # KPI Data
    kpis = [
        ["KPI", "Value"],
        ["Total Customers", len(cleaned_df)],
        ["Average Monthly Income", f"${cleaned_df['MonthlyIncome'].mean():.2f}"],
        ["Average Risk Score", f"{risk_scores_df['risk_score'].mean():.2f}"],
        ["Low Risk Customers", len(risk_scores_df[risk_scores_df['risk_category'] == 'Low Risk'])],
        ["Medium Risk Customers", len(risk_scores_df[risk_scores_df['risk_category'] == 'Medium Risk'])],
        ["High Risk Customers", len(risk_scores_df[risk_scores_df['risk_category'] == 'High Risk'])],
        ["Average Debt Ratio", f"{cleaned_df['DebtRatio'].mean():.4f}"],
        ["Average Credit Utilization", f"{cleaned_df['RevolvingUtilizationOfUnsecuredLines'].mean():.4f}"],
        ["Average Age", f"{cleaned_df['age'].mean():.1f}"]
    ]
    
    # Write KPIs
    for row_idx, row_data in enumerate(kpis, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    print("KPI sheet created")

def create_risk_summary_sheet(wb, risk_scores_df):
    """Create risk summary sheet with pivot table"""
    print("Creating risk summary sheet...")
    
    ws = wb.create_sheet("Risk Summary")
    
    # Title
    ws['A1'] = "Risk Category Distribution"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:C1')
    
    # Risk category counts
    risk_counts = risk_scores_df['risk_category'].value_counts()
    risk_percentages = risk_scores_df['risk_category'].value_counts(normalize=True) * 100
    
    # Header
    headers = ["Risk Category", "Count", "Percentage"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    for row_idx, (category, count) in enumerate(risk_counts.items(), start=4):
        ws.cell(row=row_idx, column=1, value=category)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=f"{risk_percentages[category]:.2f}%")
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Create pie chart
    pie = PieChart()
    pie.title = "Risk Category Distribution"
    
    labels = Reference(ws, min_col=1, min_row=4, max_row=len(risk_counts) + 3)
    data = Reference(ws, min_col=2, min_row=4, max_row=len(risk_counts) + 3)
    
    pie.add_data(data)
    pie.set_categories(labels)
    
    ws.add_chart(pie, "E3")
    
    print("Risk summary sheet created")

def create_income_summary_sheet(wb, cleaned_df):
    """Create income summary sheet"""
    print("Creating income summary sheet...")
    
    ws = wb.create_sheet("Income Summary")
    
    # Title
    ws['A1'] = "Income Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')
    
    # Income ranges
    income_bins = [0, 25000, 50000, 75000, 100000, float('inf')]
    income_labels = ['< 25K', '25K-50K', '50K-75K', '75K-100K', '> 100K']
    
    cleaned_df['income_range'] = pd.cut(cleaned_df['MonthlyIncome'], bins=income_bins, labels=income_labels)
    income_distribution = cleaned_df['income_range'].value_counts().sort_index()
    
    # Header
    headers = ["Income Range", "Count", "Percentage"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    total_customers = len(cleaned_df)
    for row_idx, (income_range, count) in enumerate(income_distribution.items(), start=4):
        ws.cell(row=row_idx, column=1, value=str(income_range))
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=f"{(count/total_customers)*100:.2f}%")
    
    # Statistics
    ws['A10'] = "Income Statistics"
    ws['A10'].font = Font(size=12, bold=True)
    
    stats = [
        ["Average Income", f"${cleaned_df['MonthlyIncome'].mean():.2f}"],
        ["Median Income", f"${cleaned_df['MonthlyIncome'].median():.2f}"],
        ["Min Income", f"${cleaned_df['MonthlyIncome'].min():.2f}"],
        ["Max Income", f"${cleaned_df['MonthlyIncome'].max():.2f}"]
    ]
    
    for row_idx, (stat, value) in enumerate(stats, start=11):
        ws.cell(row=row_idx, column=1, value=stat)
        ws.cell(row=row_idx, column=2, value=value)
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Create bar chart
    bar = BarChart()
    bar.title = "Income Distribution"
    bar.type = "col"
    
    data = Reference(ws, min_col=2, min_row=4, max_row=len(income_distribution) + 3)
    cats = Reference(ws, min_col=1, min_row=4, max_row=len(income_distribution) + 3)
    
    bar.add_data(data)
    bar.set_categories(cats)
    
    ws.add_chart(bar, "E3")
    
    print("Income summary sheet created")

def create_customer_details_sheet(wb, cleaned_df, risk_scores_df):
    """Create customer details sheet"""
    print("Creating customer details sheet...")
    
    ws = wb.create_sheet("Customer Details")
    
    # Merge dataframes
    if 'Unnamed: 0' in cleaned_df.columns:
        cleaned_df = cleaned_df.rename(columns={'Unnamed: 0': 'customer_id'})
    else:
        cleaned_df['customer_id'] = cleaned_df.index
    
    merged_df = cleaned_df.merge(risk_scores_df, on='customer_id')
    
    # Select relevant columns
    columns_to_keep = ['customer_id', 'age', 'MonthlyIncome', 'NumberOfDependents', 
                      'DebtRatio', 'RevolvingUtilizationOfUnsecuredLines', 
                      'risk_score', 'risk_category']
    
    display_df = merged_df[columns_to_keep].head(100)  # Limit to first 100 for performance
    
    # Write dataframe to sheet
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 30)
    
    print("Customer details sheet created")

def main():
    """Main function to generate Excel report"""
    print("\n" + "="*50)
    print("GENERATING EXCEL REPORT")
    print("="*50)
    
    # Define paths
    cleaned_data_path = '../data/cleaned_data.csv'
    risk_scores_path = '../reports/risk_scores.csv'
    output_path = '../reports/excel_report.xlsx'
    
    # Load data
    cleaned_df, risk_scores_df = load_data(cleaned_data_path, risk_scores_path)
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    create_kpi_sheet(wb, cleaned_df, risk_scores_df)
    create_risk_summary_sheet(wb, risk_scores_df)
    create_income_summary_sheet(wb, cleaned_df)
    create_customer_details_sheet(wb, cleaned_df, risk_scores_df)
    
    # Save workbook
    wb.save(output_path)
    print(f"\nExcel report saved to {output_path}")
    
    print("\n" + "="*50)
    print("EXCEL REPORT GENERATION COMPLETED")
    print("="*50)

if __name__ == "__main__":
    main()
